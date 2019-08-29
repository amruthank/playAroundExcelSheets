'''
        Test!
	Command to execute the script: python PlayAroundExcel.py existing_db.xlsx latest_component_file.xlsx

	File formats before running the script:	
		a) existing_db.xlsx : 
			Column A: Single Company Name
			Column B: Only Component Names

			Sheet names are saved with the company name.
			example: Sheet1->Honeywell
				 Sheet2->Google etc

		b)latest_component_file.xlsx:
			Column H: Only Component Names

			1st Row should be a heading!!
'''


import os 
import sys
import openpyxl


#Check for the valid files!!
if not os.path.isfile(sys.argv[1]):
	raise Exception("File ** %s ** does not exist!!\n Please check the file location!!"%sys.argv[1])
elif not os.path.isfile(sys.argv[2]):
	raise Exception("File ** %s ** does not exist!!\n Please check the file location!!"%sys.argv[2])

	
db_file = sys.argv[1]
src_file = sys.argv[2]


if db_file.split(".")[-1] != "xlsx":
	raise Exception("Expecting EXCEL file!!But you are trying to parse .%s format file!!"%db_file.split(".")[-1])
if src_file.split(".")[-1] != "xlsx":
	raise Exception("Expecting EXCEL file!!But you are trying to parse .%s format file!!"%src_file.split(".")[-1])

	
try:	
	old_db = openpyxl.load_workbook(db_file)
except Exception:
	raise Exception("Please install openpyxl package!!\nCommand to install the package - (pip install openpyxl)")
else:	
	old_sheets =  old_db.get_sheet_names()

	new_db = openpyxl.load_workbook(src_file)
	new_sheet = new_db.get_sheet_names()

	
#Function to check whether the latest component is already present in DB!!
def is_present(component,old_db_sheet):
	
	is_component_present = False
	for db_cnt in old_db_sheet['B']:
		if str(component) == str(db_cnt.value):
			is_component_present = True
			break
			
	return is_component_present
	
#Funcrion to find a matching component and if found then the st_component_file.xlsx cells are named with the Company names.
def find_existing_component(new_con_file,old_db_con):

	is_present = False
	for new_cell in (new_con_file['H'])[1:]:
			coordinate = str('H'+str(new_cell.row))
			
			if str(new_cell.value) != "" and str(new_cell.value) != "None":
				for old_db_cell in old_db_con['B']:				
					if str(new_cell.value) == str(old_db_cell.value):
						new_con_file[coordinate] = old_db_con.title
						is_present = True
	return is_present


#Functio to append the latest component to the existing_db.xlsx	
def append_db(new_con_file,old_db_con):

	for row in (new_con_file['H'])[2:]:
		if ((str(row.value) not in ["" , "None"]) and (str(row.value) != old_db_con.title)):
			if not is_present(row.value,old_db_con):
				old_coordinate = str('B'+str(old_db_con.max_row+1))
				old_db_con[old_coordinate] = str(row.value)
				
	old_db.save(db_file)	
	
def main():

	for osheet in old_sheets:
		
		new_con_file = new_db.get_sheet_by_name(new_sheet[0])
		old_db_con = old_db.get_sheet_by_name(osheet)
		
		
		try:
			is_present = find_existing_component(new_con_file,old_db_con)
		except Exception:
			raise Exception("Error while parsing the files")

		
		if is_present:
			try:
				append_db(new_con_file,old_db_con)
			except Exception:
				raise Exception("Error appending existing db!!")	

	new_db.save(src_file)
	

if __name__ == "__main__":
		main()




				
